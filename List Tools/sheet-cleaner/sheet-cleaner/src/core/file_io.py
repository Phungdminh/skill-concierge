"""File input/output for Excel and CSV.

Responsibilities:
- Detect encoding for CSV files (utf-8, utf-8-sig, cp1252, latin-1).
- Read multi-sheet Excel files; list sheet names.
- Write cleaned data to .xlsx or .csv.
- Highlight invalid cells (yellow fill) when writing Excel.
"""

from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..utils.errors import (
    MSG_ENCODING_FAILED,
    MSG_FILE_EMPTY,
    MSG_FILE_LOCKED,
    MSG_FILE_NOT_FOUND,
    MSG_FILE_PASSWORD,
    MSG_READ_FAILED,
    MSG_UNSUPPORTED_FORMAT,
    MSG_WRITE_FAILED,
    EmptyFileError,
    FileReadError,
    FileWriteError,
    LockedFileError,
    PasswordProtectedError,
    UnsupportedFormatError,
)
from ..utils.logger import get_logger


logger = get_logger("file_io")

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
CSV_ENCODINGS_TO_TRY = ["utf-8-sig", "utf-8", "cp1252", "cp1258", "latin-1"]
LARGE_FILE_BYTES = 50 * 1024 * 1024  # 50 MB

# Cell highlight colors
INVALID_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
HEADER_FONT = Font(bold=True)


def get_extension(path: str | Path) -> str:
    """Return lowercase file extension including the leading dot."""
    return Path(path).suffix.lower()


def is_supported(path: str | Path) -> bool:
    return get_extension(path) in SUPPORTED_EXTENSIONS


def file_size(path: str | Path) -> int:
    return os.path.getsize(path)


def is_large_file(path: str | Path) -> bool:
    try:
        return file_size(path) > LARGE_FILE_BYTES
    except OSError:
        return False


def list_excel_sheets(path: str | Path) -> list[str]:
    """Return list of sheet names in an Excel file."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except PermissionError as e:
        raise LockedFileError(MSG_FILE_LOCKED, str(e)) from e
    except FileNotFoundError as e:
        raise FileReadError(MSG_FILE_NOT_FOUND, str(e)) from e
    except Exception as e:
        msg = str(e).lower()
        if "password" in msg or "encrypt" in msg:
            raise PasswordProtectedError(MSG_FILE_PASSWORD, str(e)) from e
        raise FileReadError(MSG_READ_FAILED, str(e)) from e


def detect_csv_encoding(path: str | Path) -> str:
    """Try common encodings; return the first that decodes the whole file."""
    last_err: Exception | None = None
    for enc in CSV_ENCODINGS_TO_TRY:
        try:
            with open(path, "r", encoding=enc) as f:
                # Read a chunk to validate; many files are small enough to read fully
                f.read()
            logger.info("CSV encoding detected: %s for %s", enc, path)
            return enc
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except Exception as e:
            last_err = e
            continue
    logger.error("Failed to detect CSV encoding for %s: %s", path, last_err)
    raise FileReadError(MSG_ENCODING_FAILED, str(last_err) if last_err else None)


def detect_csv_delimiter(path: str | Path, encoding: str) -> str:
    """Sniff CSV delimiter from the first ~8 KB."""
    try:
        with open(path, "r", encoding=encoding, newline="") as f:
            sample = f.read(8192)
        if not sample:
            return ","
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            return dialect.delimiter
        except csv.Error:
            return ","
    except Exception:
        return ","


def read_csv(path: str | Path) -> pd.DataFrame:
    """Read a CSV file, auto-detecting encoding and delimiter."""
    encoding = detect_csv_encoding(path)
    delimiter = detect_csv_delimiter(path, encoding)
    try:
        df = pd.read_csv(
            path,
            encoding=encoding,
            sep=delimiter,
            dtype=str,
            keep_default_na=False,
            na_values=[""],
            engine="python",
            on_bad_lines="skip",
        )
    except PermissionError as e:
        raise LockedFileError(MSG_FILE_LOCKED, str(e)) from e
    except FileNotFoundError as e:
        raise FileReadError(MSG_FILE_NOT_FOUND, str(e)) from e
    except Exception as e:
        raise FileReadError(MSG_READ_FAILED, str(e)) from e

    if df.empty:
        raise EmptyFileError(MSG_FILE_EMPTY)
    return df


def read_excel(path: str | Path, sheet_name: str | None = None) -> pd.DataFrame:
    """Read an Excel file (xlsx/xls). Reads first sheet when sheet_name is None."""
    try:
        df = pd.read_excel(
            path,
            sheet_name=sheet_name if sheet_name is not None else 0,
            dtype=object,
            engine=None,
            keep_default_na=False,
            na_values=[""],
        )
    except PermissionError as e:
        raise LockedFileError(MSG_FILE_LOCKED, str(e)) from e
    except FileNotFoundError as e:
        raise FileReadError(MSG_FILE_NOT_FOUND, str(e)) from e
    except Exception as e:
        msg = str(e).lower()
        if "password" in msg or "encrypt" in msg:
            raise PasswordProtectedError(MSG_FILE_PASSWORD, str(e)) from e
        raise FileReadError(MSG_READ_FAILED, str(e)) from e

    # When sheet_name is None pandas may return a dict on some engines; guard
    if isinstance(df, dict):
        first_key = next(iter(df))
        df = df[first_key]

    if df.empty:
        raise EmptyFileError(MSG_FILE_EMPTY)

    # Convert everything to string for safe cleaning, except keep NaN
    df = df.astype(object)
    return df


def read_any(path: str | Path, sheet_name: str | None = None) -> pd.DataFrame:
    """Read any supported file format."""
    if not is_supported(path):
        raise UnsupportedFormatError(MSG_UNSUPPORTED_FORMAT)
    ext = get_extension(path)
    if ext == ".csv":
        return read_csv(path)
    return read_excel(path, sheet_name=sheet_name)


def write_excel(
    path: str | Path,
    df: pd.DataFrame,
    invalid_cells: Iterable[tuple[int, str]] | None = None,
) -> None:
    """Write DataFrame to an .xlsx file.

    Args:
        path: Destination path.
        df: DataFrame to write.
        invalid_cells: Iterable of (row_index, column_name) pairs to highlight yellow.
            row_index is the DataFrame index (0-based, before header row).
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Write header
        columns = list(df.columns)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = HEADER_FONT

        # Write data
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row[col_name]
                if pd.isna(value):
                    value = None
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Highlight invalid cells
        if invalid_cells:
            col_to_idx = {name: i + 1 for i, name in enumerate(columns)}
            for df_row_idx, col_name in invalid_cells:
                if col_name not in col_to_idx:
                    continue
                excel_row = df_row_idx + 2  # +1 for 1-based, +1 for header
                excel_col = col_to_idx[col_name]
                ws.cell(row=excel_row, column=excel_col).fill = INVALID_FILL

        # Auto-size columns (approximate)
        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            try:
                series = df[col_name].astype(str)
                max_len = max(series.map(len).max(), len(str(col_name)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
            except Exception:
                ws.column_dimensions[col_letter].width = 18

        wb.save(path)
        logger.info("Wrote Excel file: %s (%d rows)", path, len(df))
    except PermissionError as e:
        raise FileWriteError(MSG_FILE_LOCKED, str(e)) from e
    except Exception as e:
        raise FileWriteError(MSG_WRITE_FAILED, str(e)) from e


def write_csv(path: str | Path, df: pd.DataFrame) -> None:
    """Write DataFrame to a CSV file (utf-8-sig for Excel compatibility)."""
    try:
        df.to_csv(path, index=False, encoding="utf-8-sig")
        logger.info("Wrote CSV file: %s (%d rows)", path, len(df))
    except PermissionError as e:
        raise FileWriteError(MSG_FILE_LOCKED, str(e)) from e
    except Exception as e:
        raise FileWriteError(MSG_WRITE_FAILED, str(e)) from e


def write_any(
    path: str | Path,
    df: pd.DataFrame,
    invalid_cells: Iterable[tuple[int, str]] | None = None,
) -> None:
    """Write DataFrame to whichever format the extension implies."""
    ext = get_extension(path)
    if ext == ".csv":
        write_csv(path, df)
    elif ext in (".xlsx", ".xls"):
        # Always write as xlsx even if user picked .xls
        out = Path(path)
        if ext == ".xls":
            out = out.with_suffix(".xlsx")
        write_excel(out, df, invalid_cells=invalid_cells)
    else:
        raise UnsupportedFormatError(MSG_UNSUPPORTED_FORMAT)
